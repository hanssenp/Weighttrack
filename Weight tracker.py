# -*- coding: utf-8 -*-
"""
Created on Thu May  3 09:44:11 2018

@author: peter
"""

# libraries
import time
import datetime
import numpy as np
import openpyxl
import csv
import matplotlib.pyplot as plt

def main():
    # start time is 4am Sun, Apr 29 2018
    time.ctime(1524988800.0398614)
    tstart = 1524988800.0398614

    # number of weeks elapsed; also equal to current week number
    tcurrent = time.time()
    weekscurrent = round((tcurrent - tstart)/604800)

    # Today's weigh in
    wcurrent = float(input("Enter today's weight in pounds: "))

    # save workbook as xcel document

    # open workbook
    from openpyxl import Workbook
    wb = Workbook()
    openpyxl.load_workbook('weight.xlsx', data_only=True) # might need .xltx ?
    Workbook(write_only=True)
    ws = wb.active # worksheet
    ws.title = "Weighins"
    ws1 = wb["Weighins"]
    # account for spreadsheet row number disparity (cell A2 is actually week 1,
    # because A1 is the column name)
    rowno = weekscurrent + 1 # Excel row number
    ws1.cell(row=rowno, column=3).value=wcurrent # write current weight in appropriate cell

    # save workbook as xcel document
    wb.save('weight2.xlsx')

# from openpyxl import load_workbook
# wb2 = load_workbook('test.xlsx')

col_range = ws['A:C'] # first three columns in use
row_range = ws[1:53] # 52 weeks in a year
colweek = ws['A'] # "week" column
rowcurr = ws[rowno] # row for current week's value

# edit line with csv
with open("weightcomma.csv", mode='w', newline=' ') as csvfile:
    writer = csv.writer(csvfile, delimiter=',')
    writer.writerow([rowno])
    writer.writerow()
    
#csvfile = csv.reader(open("weightcomma.csv", mode='r'))
#lines = list(csvfile)
#lines[rowno][3] = wcurrent
#writer = csv.writer(open("weightcomma.csv", mode='w')) 
#writer.writerows(lines)   

# Try again

# Functional! This section will get the csv as a list, then change the current
# week's value to the current weight as input above.
f1 = open("weightcomma.csv", "r")
csvfile = csv.reader(f1)
lines = list(csvfile)

# You are getting the csv file here as a list of lists. So:
# Select the row you want -- current week -- as a list, taken from the larger
# list of all the rows in the csv. Delete the third entry on that small list,
# then add a new third entry with current weigh -- effectively, replacing the
# third (empty) cell with the correct value for current weight. Then
# replace (delete and insert) the appropriate row in the big list with your new, modified
# small list. Then write the updated list to the csv.
lines2=lines[weekscurrent].copy()
lines2.pop(2)
lines2.insert(2, wcurrent)
lines.pop(weekscurrent)
lines.insert(weekscurrent,lines2)

# Close the original file, we're done with it
f1.close()

# Attempt to write this list back into the csv.
# Currently writing to a seperate csv, "weightcomma2.csv," for test purposes
# So that you don't have to manually remake "weightcomma" every time.
# writer = csv.writer(open("weightcomma2.csv", mode='w'))
# writer.writerows(lines)

# Some online have suggested a "with" loop:
with open("weightcomma2.csv", "w") as f:
    writer = csv.writer(f)
    writer.writerows(lines)



# Generate date range
from datetime import date, datetime, timedelta
def perdelta(start, end, delta):
    curr = start
    while curr < end:
        yield curr
        curr += delta
        
    # timeframe = np.array()
for result in perdelta(date(2018, 4, 29), date(2019, 4, 29), timedelta(days=7)):
    print(result)




wb.save('weight2.xlsx') # save changes

# set graph x = week number to display data over time; set y = weight at each week
x = ws['A2':'A53']
y = ws['C2':'C53']






# save workbook as xcel document
wb = openpyxl.load_workbook('weight.xlsx') # might need .xltx ?
wb.template = False
wb.save('weight2.xlsx')


# test read csv file
import csv
with open('weightcomma.csv', newline='') as csvfile:
    weightreader = csv.reader(csvfile, delimiter=' ', quotechar='|')
    for row in weightreader:
        print(', '.join(row))

import csv
with open('eggs.csv', 'w', newline='') as csvfile:
    weightwriter = csv.writer(csvfile, delimiter=' ',
                            quotechar='|', quoting=csv.QUOTE_MINIMAL)
    weightwriter.writerow([weekscurrent] * 5 + ['Baked Beans'])
    weightwriter.writerow(['Spam', 'Lovely Spam', 'Wonderful Spam'])



# Enter today's weigh in in excel
open('weightcomma.csv', 'w', newline='') as csvfile:
    weightwriter = csv.writer(csvfile, delimiter=' ',
                            quotechar='|', quoting=csv.QUOTE_MINIMAL)
    spamwriter.writerow(['Spam'] * 5 + ['Baked Beans'])
    spamwriter.writerow(['Spam', 'Lovely Spam', 'Wonderful Spam'])

x = np.linspace (0,100000)
weight0 = 160
tsec = time.ctime()
tcurrent = tsec - 
tcurrent = int(input("Week number?"))
wcurrent = float(input("Weight today?"))
plt.axhline(y=160, color='r', linestyle='-')
plt.axis([0, 4, 0, 170])
plt.plot([tcurrent], [wcurrent], 'bo')
plt.show()
# weight0 = y=160
# weightdes = y=145
 